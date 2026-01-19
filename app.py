# -*- coding: utf-8 -*-
"""
邮件批量发送系统 - Web版（无pandas依赖）
"""

import os
import smtplib
from flask import Flask, request, jsonify, session
from werkzeug.utils import secure_filename
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header
from email.utils import formataddr
import mimetypes
from openpyxl import load_workbook
import zipfile

app = Flask(__name__)
app.secret_key = 'email_sender_secret_key_2024'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'attachments'), exist_ok=True)

DEFAULT_CONFIG = {
    'smtp_server': 'smtp.exmail.qq.com',
    'smtp_port': 465
}

HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>邮件批量发送系统</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Microsoft YaHei', sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; padding: 20px; }
        .container { max-width: 1400px; margin: 0 auto; background: white; border-radius: 15px; box-shadow: 0 20px 60px rgba(0,0,0,0.3); overflow: hidden; }
        .header { background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%); color: white; padding: 25px 30px; text-align: center; }
        .header h1 { font-size: 28px; margin-bottom: 5px; }
        .header p { opacity: 0.8; font-size: 14px; }
        .main-content { display: flex; min-height: 600px; flex-wrap: wrap; }
        .left-panel { width: 45%; padding: 25px; border-right: 1px solid #eee; min-width: 300px; }
        .right-panel { width: 55%; padding: 25px; background: #fafafa; min-width: 300px; flex: 1; }
        .section { margin-bottom: 25px; }
        .section-title { font-size: 16px; font-weight: bold; color: #333; margin-bottom: 15px; padding-bottom: 10px; border-bottom: 2px solid #667eea; }
        .btn { padding: 12px 24px; border: none; border-radius: 8px; cursor: pointer; font-size: 14px; font-weight: bold; transition: all 0.3s; display: inline-flex; align-items: center; gap: 8px; margin: 5px; }
        .btn-primary { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; }
        .btn-primary:hover { transform: translateY(-2px); box-shadow: 0 5px 20px rgba(102, 126, 234, 0.4); }
        .btn-success { background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); color: white; }
        .btn-success:hover { transform: translateY(-2px); box-shadow: 0 5px 20px rgba(17, 153, 142, 0.4); }
        .btn-warning { background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); color: white; }
        .btn:disabled { opacity: 0.5; cursor: not-allowed; transform: none !important; }
        .file-input { display: none; }
        .upload-area { display: flex; gap: 10px; flex-wrap: wrap; }
        .status-box { background: #e8f5e9; border: 1px solid #a5d6a7; border-radius: 8px; padding: 15px; margin-top: 15px; }
        .company-table { width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 13px; }
        .company-table th, .company-table td { padding: 10px; text-align: left; border-bottom: 1px solid #eee; }
        .company-table th { background: #f5f5f5; font-weight: bold; color: #555; }
        .company-table tr:hover { background: #f0f7ff; }
        .company-table tr.active { background: #e3f2fd; }
        .company-table tr.sent { background: #e8f5e9; }
        .table-container { max-height: 250px; overflow-y: auto; border: 1px solid #ddd; border-radius: 8px; }
        .template-input { width: 100%; padding: 12px; border: 1px solid #ddd; border-radius: 8px; font-size: 14px; margin-bottom: 10px; }
        .template-textarea { width: 100%; height: 150px; padding: 12px; border: 1px solid #ddd; border-radius: 8px; font-size: 14px; resize: vertical; font-family: inherit; }
        .preview-box { background: white; border: 1px solid #ddd; border-radius: 8px; padding: 15px; margin-top: 15px; }
        .preview-recipient { font-size: 14px; font-weight: bold; color: #1976d2; margin-bottom: 10px; padding-bottom: 10px; border-bottom: 1px solid #eee; }
        .preview-content { white-space: pre-wrap; font-size: 13px; line-height: 1.6; color: #333; max-height: 200px; overflow-y: auto; }
        .nav-buttons { display: flex; gap: 10px; margin-top: 15px; justify-content: center; flex-wrap: wrap; }
        .btn-nav { padding: 8px 16px; background: #f5f5f5; border: 1px solid #ddd; border-radius: 8px; cursor: pointer; font-size: 13px; }
        .btn-nav:hover:not(:disabled) { background: #e0e0e0; }
        .progress-bar { width: 100%; height: 8px; background: #e0e0e0; border-radius: 4px; overflow: hidden; margin: 10px 0; }
        .progress-fill { height: 100%; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); transition: width 0.3s; }
        .progress-text { text-align: center; font-size: 13px; color: #666; }
        .send-section { text-align: center; padding: 15px; background: white; border-radius: 8px; margin-top: 15px; }
        .btn-send { padding: 15px 40px; font-size: 16px; }
        .hint { font-size: 11px; color: #999; margin-top: 8px; }
        .attachment-list { margin-top: 10px; padding: 10px; background: #f5f5f5; border-radius: 5px; font-size: 12px; }
        .loading { display: none; text-align: center; padding: 20px; }
        .loading.show { display: block; }
        .spinner { border: 3px solid #f3f3f3; border-top: 3px solid #667eea; border-radius: 50%; width: 30px; height: 30px; animation: spin 1s linear infinite; margin: 0 auto 10px; }
        @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
        .toast { position: fixed; top: 20px; right: 20px; padding: 15px 25px; border-radius: 8px; color: white; font-weight: bold; z-index: 1000; animation: slideIn 0.3s ease; }
        .toast.success { background: #4caf50; }
        .toast.error { background: #f44336; }
        @keyframes slideIn { from { transform: translateX(100%); opacity: 0; } to { transform: translateX(0); opacity: 1; } }
        @media (max-width: 800px) { .left-panel, .right-panel { width: 100%; } }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📧 邮件批量发送系统</h1>
            <p>导入Excel → 编辑模板 → 预览确认 → 发送</p>
        </div>
        <div class="main-content">
            <div class="left-panel">
                <div class="section">
                    <div class="section-title">1. 导入数据</div>
                    <div class="upload-area">
                        <input type="file" id="excelFile" class="file-input" accept=".xlsx">
                        <button class="btn btn-primary" onclick="document.getElementById('excelFile').click()">📂 导入Excel</button>
                        <input type="file" id="attachFiles" class="file-input" multiple>
                        <button class="btn btn-warning" onclick="document.getElementById('attachFiles').click()">📎 添加附件</button>
                    </div>
                    <div id="importStatus" class="status-box" style="display:none;"></div>
                    <div id="attachStatus" class="attachment-list" style="display:none;"></div>
                </div>
                <div class="section">
                    <div class="section-title">2. 公司列表</div>
                    <div class="table-container">
                        <table class="company-table"><thead><tr><th>#</th><th>公司</th><th>负责人</th><th>邮箱</th></tr></thead>
                        <tbody id="companyList"><tr><td colspan="4" style="text-align:center;color:#999;padding:20px;">请先导入Excel</td></tr></tbody></table>
                    </div>
                </div>
                <div class="section">
                    <div class="section-title">3. 发件人设置</div>
                    <input type="text" id="senderName" class="template-input" placeholder="发件人姓名">
                    <input type="email" id="senderEmail" class="template-input" placeholder="发件人邮箱">
                    <input type="password" id="senderPassword" class="template-input" placeholder="邮箱授权码">
                </div>
            </div>
            <div class="right-panel">
                <div class="section">
                    <div class="section-title">4. 邮件模板</div>
                    <input type="text" id="emailSubject" class="template-input" placeholder="邮件主题" value="合作邀请函">
                    <textarea id="emailTemplate" class="template-textarea">尊敬的{company_name}的{contact_person}：

您好！

感谢您百忙之中阅读此邮件。我们诚挚地希望能与贵公司建立合作关系。

如有任何问题，欢迎随时联系。

祝好！
{sender_name}</textarea>
                    <div class="hint">变量: {company_name}=公司名, {contact_person}=负责人, {sender_name}=你的名字</div>
                </div>
                <div class="section">
                    <div class="section-title">5. 预览与发送</div>
                    <div class="progress-bar"><div class="progress-fill" id="progressFill" style="width:0%"></div></div>
                    <div class="progress-text" id="progressText">进度: 0/0</div>
                    <div class="preview-box">
                        <div class="preview-recipient" id="previewRecipient">请先导入Excel</div>
                        <div class="preview-content" id="previewContent"></div>
                    </div>
                    <div class="nav-buttons">
                        <button class="btn-nav" id="btnPrev" onclick="prevCompany()" disabled>◀ 上一个</button>
                        <button class="btn-nav" onclick="refreshPreview()">🔄 刷新</button>
                        <button class="btn-nav" id="btnNext" onclick="nextCompany()" disabled>下一个 ▶</button>
                    </div>
                </div>
                <div class="send-section">
                    <div style="margin-bottom:10px;">
                        <label><input type="radio" name="sendMode" value="single" checked onchange="updateSendMode()"> 单个发送</label>
                        <label style="margin-left:15px;"><input type="radio" name="sendMode" value="batch" onchange="updateSendMode()"> 批量发送</label>
                    </div>
                    <div id="singleSendArea">
                        <button class="btn btn-success btn-send" id="btnSend" onclick="sendEmail()" disabled>✉️ 发送当前邮件</button>
                    </div>
                    <div id="batchSendArea" style="display:none;">
                        <button class="btn btn-success btn-send" id="btnSendAll" onclick="sendAllEmails()" disabled style="background:linear-gradient(135deg,#ff6b6b 0%,#ee5a24 100%);">🚀 一键发送全部</button>
                    </div>
                </div>
                <div class="loading" id="loading"><div class="spinner"></div><div>发送中...</div></div>
            </div>
        </div>
    </div>
    <script>
        let companies=[], currentIndex=0, sentStatus=[];
        document.getElementById('excelFile').addEventListener('change', function(e) {
            const file=e.target.files[0]; if(!file) return;
            const formData=new FormData(); formData.append('file', file);
            fetch('/upload_excel', {method:'POST', body:formData}).then(r=>r.json()).then(data=>{
                if(data.success) { companies=data.companies; sentStatus=new Array(companies.length).fill(false); currentIndex=0;
                    document.getElementById('importStatus').style.display='block';
                    document.getElementById('importStatus').innerHTML='✓ 导入 <b>'+data.count+'</b> 个公司';
                    renderTable(); updateUI(); showToast('导入成功','success');
                } else showToast(data.error,'error');
            }).catch(err=>showToast('上传失败','error'));
        });
        document.getElementById('attachFiles').addEventListener('change', function(e) {
            const files=e.target.files; if(!files.length) return;
            const formData=new FormData(); for(let f of files) formData.append('files', f);
            fetch('/upload_attachments', {method:'POST', body:formData}).then(r=>r.json()).then(data=>{
                if(data.success) { document.getElementById('attachStatus').style.display='block';
                    document.getElementById('attachStatus').innerHTML='📎 '+data.files.join(', ');
                    showToast('添加'+data.count+'个附件','success');
                } else showToast(data.error,'error');
            }).catch(err=>showToast('上传失败','error'));
        });
        function renderTable() {
            const tbody=document.getElementById('companyList');
            if(!companies.length) { tbody.innerHTML='<tr><td colspan="4" style="text-align:center;color:#999;padding:20px;">请先导入Excel</td></tr>'; return; }
            tbody.innerHTML=companies.map((c,i)=>'<tr class="'+(i===currentIndex?'active':'')+' '+(sentStatus[i]?'sent':'')+'" onclick="selectCompany('+i+')" style="cursor:pointer"><td>'+(i+1)+'</td><td>'+c.name+'</td><td>'+c.contact+'</td><td>'+c.email+'</td></tr>').join('');
        }
        function selectCompany(i) { currentIndex=i; updateUI(); }
        function prevCompany() { if(currentIndex>0) { currentIndex--; updateUI(); } }
        function nextCompany() { if(currentIndex<companies.length-1) { currentIndex++; updateUI(); } }
        function updateUI() {
            renderTable();
            document.getElementById('btnPrev').disabled=currentIndex<=0;
            document.getElementById('btnNext').disabled=currentIndex>=companies.length-1;
            document.getElementById('btnSend').disabled=!companies.length;
            document.getElementById('btnSendAll').disabled=!companies.length;
            const p=companies.length?((currentIndex+1)/companies.length*100):0;
            document.getElementById('progressFill').style.width=p+'%';
            document.getElementById('progressText').textContent='进度: '+(currentIndex+1)+'/'+companies.length+' | 已发送: '+sentStatus.filter(s=>s).length;
            document.getElementById('btnSend').textContent=sentStatus[currentIndex]?'✓ 已发送':'✉️ 发送当前邮件';
            document.getElementById('btnSend').style.background=sentStatus[currentIndex]?'#9e9e9e':'';
            refreshPreview();
        }
        function refreshPreview() {
            if(!companies.length) { document.getElementById('previewRecipient').textContent='请先导入Excel'; document.getElementById('previewContent').textContent=''; return; }
            const c=companies[currentIndex], tpl=document.getElementById('emailTemplate').value, name=document.getElementById('senderName').value;
            document.getElementById('previewRecipient').textContent='收件人: '+c.name+' - '+c.contact+' <'+c.email+'>';
            try { document.getElementById('previewContent').textContent='【主题】'+document.getElementById('emailSubject').value+'\\n\\n'+tpl.replace(/{company_name}/g,c.name).replace(/{contact_person}/g,c.contact).replace(/{sender_name}/g,name); }
            catch(e) { document.getElementById('previewContent').textContent='模板错误'; }
        }
        function sendEmail() {
            if(!companies.length) return;
            const c=companies[currentIndex];
            if(!confirm('发送邮件给:\\n'+c.name+' <'+c.email+'>?')) return;
            document.getElementById('loading').classList.add('show');
            document.getElementById('btnSend').disabled=true;
            fetch('/send_email', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({index:currentIndex, template:document.getElementById('emailTemplate').value, subject:document.getElementById('emailSubject').value, sender_name:document.getElementById('senderName').value, sender_email:document.getElementById('senderEmail').value, sender_password:document.getElementById('senderPassword').value})})
            .then(r=>r.json()).then(data=>{
                document.getElementById('loading').classList.remove('show');
                if(data.success) { sentStatus[currentIndex]=true; showToast(data.message,'success'); if(currentIndex<companies.length-1) currentIndex++; updateUI(); }
                else { showToast(data.error,'error'); document.getElementById('btnSend').disabled=false; }
            }).catch(err=>{ document.getElementById('loading').classList.remove('show'); showToast('发送失败','error'); document.getElementById('btnSend').disabled=false; });
        }
        function sendAllEmails() {
            if(!companies.length) return;
            if(!confirm('一次性发送给所有 '+companies.length+' 个公司?')) return;
            document.getElementById('loading').classList.add('show');
            document.getElementById('btnSendAll').disabled=true;
            fetch('/send_all', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({template:document.getElementById('emailTemplate').value, subject:document.getElementById('emailSubject').value, sender_name:document.getElementById('senderName').value, sender_email:document.getElementById('senderEmail').value, sender_password:document.getElementById('senderPassword').value})})
            .then(r=>r.json()).then(data=>{
                document.getElementById('loading').classList.remove('show');
                if(data.success) { data.results.forEach(r=>{ if(r.success) sentStatus[r.index]=true; }); updateUI(); showToast('成功:'+data.success_count+' 失败:'+data.fail_count, data.fail_count?'error':'success'); }
                else showToast(data.error,'error');
                document.getElementById('btnSendAll').disabled=false;
            }).catch(err=>{ document.getElementById('loading').classList.remove('show'); showToast('发送失败','error'); document.getElementById('btnSendAll').disabled=false; });
        }
        function showToast(msg,type) { const t=document.createElement('div'); t.className='toast '+type; t.textContent=msg; document.body.appendChild(t); setTimeout(()=>t.remove(),3000); }
        function updateSendMode() { const m=document.querySelector('input[name="sendMode"]:checked').value; document.getElementById('singleSendArea').style.display=m==='single'?'block':'none'; document.getElementById('batchSendArea').style.display=m==='batch'?'block':'none'; }
        document.getElementById('emailTemplate').addEventListener('input', refreshPreview);
        document.getElementById('emailSubject').addEventListener('input', refreshPreview);
        document.getElementById('senderName').addEventListener('input', refreshPreview);
    </script>
</body>
</html>'''


@app.route('/')
def index():
    return HTML_TEMPLATE


@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有选择文件'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '没有选择文件'})
    try:
        original_name = file.filename
        lower_name = (original_name or '').lower()
        if not lower_name.endswith('.xlsx'):
            return jsonify({
                'success': False,
                'error': '请上传 .xlsx 格式的Excel文件（openpyxl不支持 .xls）。请用Excel打开后“另存为 -> Excel 工作簿(*.xlsx)”再上传。'
            })

        filename = secure_filename(file.filename) or 'upload.xlsx'
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        try:
            wb = load_workbook(filepath)
        except Exception as e:
            return jsonify({
                'success': False,
                'error': f'Excel解析失败：{str(e)}。请确认文件能在Excel里正常打开，并重新"另存为 .xlsx"后再上传。'
            })
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        if '公司名称' not in headers or '邮箱地址' not in headers:
            return jsonify({'success': False, 'error': 'Excel必须包含"公司名称"和"邮箱地址"列'})
        
        name_idx = headers.index('公司名称')
        email_idx = headers.index('邮箱地址')
        contact_idx = headers.index('负责人') if '负责人' in headers else -1
        
        companies = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[name_idx] or '').strip()
            email = str(row[email_idx] or '').strip()
            contact = str(row[contact_idx] or '负责人').strip() if contact_idx >= 0 else '负责人'
            if name and email and name != 'None':
                companies.append({'name': name, 'email': email, 'contact': contact if contact and contact != 'None' else '负责人'})
        
        session['companies'] = companies
        session['sent_status'] = [False] * len(companies)
        return jsonify({'success': True, 'count': len(companies), 'companies': companies})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/upload_attachments', methods=['POST'])
def upload_attachments():
    if 'files' not in request.files:
        return jsonify({'success': False, 'error': '没有选择文件'})
    files = request.files.getlist('files')
    if not files or files[0].filename == '':
        return jsonify({'success': False, 'error': '没有选择文件'})
    try:
        attach_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'attachments')
        for f in os.listdir(attach_folder):
            os.remove(os.path.join(attach_folder, f))
        saved_files = []
        for file in files:
            filename = secure_filename(file.filename) or file.filename
            filepath = os.path.join(attach_folder, filename)
            file.save(filepath)
            saved_files.append({'name': filename, 'path': filepath})
        session['attachments'] = saved_files
        return jsonify({'success': True, 'count': len(saved_files), 'files': [f['name'] for f in saved_files]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/send_email', methods=['POST'])
def send_email():
    data = request.json
    index = data.get('index', 0)
    template = data.get('template', '')
    subject = data.get('subject', '')
    sender_email = data.get('sender_email', '')
    sender_password = data.get('sender_password', '')
    sender_name = data.get('sender_name', '')
    
    if not sender_email or not sender_password:
        return jsonify({'success': False, 'error': '请填写发件人邮箱和授权码'})
    
    companies = session.get('companies', [])
    if not companies or index >= len(companies):
        return jsonify({'success': False, 'error': '没有公司数据'})
    
    company = companies[index]
    try:
        msg = MIMEMultipart()
        msg['From'] = formataddr((sender_name, sender_email))
        msg['To'] = formataddr((company['name'], company['email']))
        msg['Subject'] = Header(subject, 'utf-8')
        body = template.replace('{company_name}', company['name']).replace('{contact_person}', company['contact']).replace('{sender_name}', sender_name)
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        for attach in session.get('attachments', []):
            filepath = attach['path']
            if os.path.exists(filepath):
                with open(filepath, 'rb') as f:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', 'attachment', filename=Header(attach['name'], 'utf-8').encode())
                msg.attach(part)
        
        server = smtplib.SMTP_SSL(DEFAULT_CONFIG['smtp_server'], DEFAULT_CONFIG['smtp_port'])
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()
        
        sent_status = session.get('sent_status', [False] * len(companies))
        sent_status[index] = True
        session['sent_status'] = sent_status
        return jsonify({'success': True, 'message': '发送成功: ' + company['name']})
    except smtplib.SMTPAuthenticationError:
        return jsonify({'success': False, 'error': '邮箱认证失败'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/send_all', methods=['POST'])
def send_all():
    data = request.json
    template = data.get('template', '')
    subject = data.get('subject', '')
    sender_email = data.get('sender_email', '')
    sender_password = data.get('sender_password', '')
    sender_name = data.get('sender_name', '')
    
    if not sender_email or not sender_password:
        return jsonify({'success': False, 'error': '请填写发件人邮箱和授权码'})
    
    companies = session.get('companies', [])
    if not companies:
        return jsonify({'success': False, 'error': '没有公司数据'})
    
    sent_status = session.get('sent_status', [False] * len(companies))
    results, success_count, fail_count = [], 0, 0
    
    for index, company in enumerate(companies):
        if sent_status[index]:
            results.append({'index': index, 'success': True, 'message': '已跳过'})
            continue
        try:
            msg = MIMEMultipart()
            msg['From'] = formataddr((sender_name, sender_email))
            msg['To'] = formataddr((company['name'], company['email']))
            msg['Subject'] = Header(subject, 'utf-8')
            body = template.replace('{company_name}', company['name']).replace('{contact_person}', company['contact']).replace('{sender_name}', sender_name)
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
            
            for attach in session.get('attachments', []):
                filepath = attach['path']
                if os.path.exists(filepath):
                    with open(filepath, 'rb') as f:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(f.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Disposition', 'attachment', filename=Header(attach['name'], 'utf-8').encode())
                    msg.attach(part)
            
            server = smtplib.SMTP_SSL(DEFAULT_CONFIG['smtp_server'], DEFAULT_CONFIG['smtp_port'])
            server.login(sender_email, sender_password)
            server.send_message(msg)
            server.quit()
            
            sent_status[index] = True
            success_count += 1
            results.append({'index': index, 'success': True, 'message': '成功'})
        except Exception as e:
            fail_count += 1
            results.append({'index': index, 'success': False, 'message': str(e)})
    
    session['sent_status'] = sent_status
    return jsonify({'success': True, 'total': len(companies), 'success_count': success_count, 'fail_count': fail_count, 'results': results})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
